# -*- coding: utf-8 -*-
from config import *
import sys, re, io, os, json, math, time, warnings
from datetime import datetime
from difflib import SequenceMatcher
import requests
import numpy as np
import pandas as pd
from bs4 import BeautifulSoup
warnings.filterwarnings("ignore")

def export_xlsx(results: list, abc: dict,
                week_id: str, mem=None) -> str:
    """
    3 sayfalık okunabilir xlsx raporu oluştur:
      Sayfa 1 — Simülasyon sonuçları (renkli)
      Sayfa 2 — A/B/C kupon planları
      Sayfa 3 — Öğrenme hafızası özeti
    """
    try:
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                     Border, Side, numbers)
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  openpyxl yok — pip install openpyxl")
        return ""

    wb = openpyxl.Workbook()

    # ── Renk paleti ──────────────────────────────────────────
    C = {
        "banko":  "1E8449",   # koyu yeşil
        "tek":    "2E86C1",   # mavi
        "cift":   "D4AC0D",   # sarı
        "kaos":   "C0392B",   # kırmızı
        "header": "1A252F",   # koyu lacivert
        "sub":    "2C3E50",   # koyu gri
        "light":  "EBF5FB",   # açık mavi
        "white":  "FFFFFF",
        "gray":   "BDC3C7",
        "green":  "EAFAF1",
        "yellow": "FEF9E7",
        "red":    "FDEDEC",
    }

    def _fill(hex_col):
        return PatternFill("solid", fgColor=hex_col)

    def _font(bold=False, white=False, size=11):
        return Font(bold=bold, size=size,
                    color="FFFFFF" if white else "000000")

    def _border():
        s = Side(style="thin", color="BDC3C7")
        return Border(left=s, right=s, top=s, bottom=s)

    def _center():
        return Alignment(horizontal="center", vertical="center",
                         wrap_text=True)

    def _left():
        return Alignment(horizontal="left", vertical="center",
                         wrap_text=True)

    def _hdr(ws, row, col, text, bg="header", white=True, bold=True, size=11):
        c = ws.cell(row=row, column=col, value=text)
        c.fill      = _fill(C[bg])
        c.font      = Font(bold=bold, size=size,
                           color="FFFFFF" if white else "000000")
        c.alignment = _center()
        c.border    = _border()
        return c

    # ════════════════════════════════════════════════════════
    # SAYFA 1 — SİMÜLASYON SONUÇLARI
    # ════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Simülasyon"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 4
    ws1.column_dimensions["B"].width = 28
    ws1.column_dimensions["C"].width = 8
    ws1.column_dimensions["D"].width = 8
    ws1.column_dimensions["E"].width = 7
    ws1.column_dimensions["F"].width = 7
    ws1.column_dimensions["G"].width = 7
    ws1.column_dimensions["H"].width = 14
    ws1.column_dimensions["I"].width = 5
    ws1.column_dimensions["J"].width = 12
    ws1.column_dimensions["K"].width = 7
    ws1.column_dimensions["L"].width = 5

    # Başlık
    ws1.merge_cells("A1:L1")
    t = ws1["A1"]
    t.value     = f"SPOR TOTO — AUGUR ENGINE  |  {week_id}  |  {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    t.fill      = _fill(C["header"])
    t.font      = Font(bold=True, size=13, color="FFFFFF")
    t.alignment = _center()
    ws1.row_dimensions[1].height = 28

    # Sütun başlıkları
    headers = ["#", "MAÇ", "λEv", "λDep", "1%", "X%", "2%", "ÖNERİ", "x", "BAĞLAM", "SONUÇ", "✓/✗"]
    for ci, h in enumerate(headers, 1):
        _hdr(ws1, 2, ci, h, bg="sub")
    ws1.row_dimensions[2].height = 20

    # Satırlar
    lbl_color = {
        "BANKO": C["banko"],
        "TEK":   C["tek"],
        "CIFT":  C["cift"],
        "KAOS":  C["kaos"],
    }

    for ri, r in enumerate(results, 3):
        lbl_key = r["oneri"].strip().split()[0]
        row_bg  = {
            "BANKO": C["green"],
            "TEK":   C["light"],
            "CIFT":  C["yellow"],
            "KAOS":  C["red"],
        }.get(lbl_key, C["white"])

        vals = [
            r["no"],
            r["mac"],
            r["lH"],
            r["lA"],
            r["P1"],
            r["PX"],
            r["P2"],
            r["oneri"].strip(),
            r["mul"],
            r.get("ctx",""),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws1.cell(row=ri, column=ci, value=v)
            cell.fill      = _fill(row_bg)
            cell.border    = _border()
            cell.alignment = _center() if ci != 2 else _left()

        # Öneri sütunu renkli
        oneri_cell = ws1.cell(row=ri, column=8)
        oneri_cell.fill = _fill(lbl_color.get(lbl_key, C["gray"]))
        oneri_cell.font = Font(bold=True, size=11, color="FFFFFF")
        ws1.row_dimensions[ri].height = 18

    # Özet satırı
    last = len(results) + 3
    ws1.merge_cells(f"A{last}:L{last}")
    banko = sum(1 for r in results if r["oneri"].startswith("BANKO"))
    tek   = sum(1 for r in results if r["oneri"].startswith("TEK"))
    cift  = sum(1 for r in results if r["oneri"].startswith("CIFT"))
    kaos  = sum(1 for r in results if r["oneri"].startswith("KAOS"))
    total_cols = 1
    for r in results: total_cols *= r["mul"]
    summary = ws1[f"A{last}"]
    summary.value     = (f"Banko: {banko}  |  Tek: {tek}  |  "
                         f"Cift: {cift}  |  Kaos: {kaos}  |  "
                         f"Toplam: {total_cols:,} kolon = {total_cols*10:,} TL")
    summary.fill      = _fill(C["sub"])
    summary.font      = Font(bold=True, size=11, color="FFFFFF")
    summary.alignment = _center()
    ws1.row_dimensions[last].height = 22

    # ════════════════════════════════════════════════════════
    # SAYFA 2 — KUPON PLANLARI (A / B / C)
    # ════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Kupon Planları")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 4
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 14

    ws2.merge_cells("A1:E1")
    t2 = ws2["A1"]
    t2.value     = "KUPON PLANLARI — A / B / C"
    t2.fill      = _fill(C["header"])
    t2.font      = Font(bold=True, size=13, color="FFFFFF")
    t2.alignment = _center()
    ws2.row_dimensions[1].height = 28

    # Başlık satırı
    _hdr(ws2, 2, 1, "#", bg="sub")
    _hdr(ws2, 2, 2, "MAÇ", bg="sub")
    for ci, (letter, d) in enumerate(abc.items(), 3):
        _hdr(ws2, 2, ci,
             f"KUPON {letter}\n{d['budget']:,} TL\n→ {d['cols']} kol = {d['cost']:,} TL",
             bg="sub")
        ws2.column_dimensions[get_column_letter(ci)].width = 16
    ws2.row_dimensions[2].height = 45

    for ri, r in enumerate(results, 3):
        ws2.cell(row=ri, column=1, value=r["no"]).border = _border()
        c = ws2.cell(row=ri, column=2, value=r["mac"])
        c.border    = _border()
        c.alignment = _left()

        for ci, (letter, d) in enumerate(abc.items(), 3):
            opt_r = next((x for x in d["results"] if x["no"] == r["no"]), r)
            lbl   = opt_r["oneri"].strip()
            lbl_k = lbl.split()[0]
            bg    = lbl_color.get(lbl_k, C["gray"])
            cell  = ws2.cell(row=ri, column=ci, value=lbl)
            cell.fill      = _fill(bg)
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = _center()
            cell.border    = _border()
        ws2.row_dimensions[ri].height = 18

    # Özet satırları
    last2 = len(results) + 3
    for ci, (letter, d) in enumerate(abc.items(), 3):
        cell = ws2.cell(row=last2, column=ci,
                        value=f"{d['cols']} kolon\n{d['cost']:,} TL")
        cell.fill      = _fill(C["sub"])
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = _center()
        cell.border    = _border()
    ws2.row_dimensions[last2].height = 32

    # ════════════════════════════════════════════════════════
    # SAYFA 3 — ÖĞRENME HAFIZASI
    # ════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Hafıza")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 16
    ws3.column_dimensions["C"].width = 16

    ws3.merge_cells("A1:C1")
    t3 = ws3["A1"]
    t3.value     = "ÖĞRENME HAFIZASI"
    t3.fill      = _fill(C["header"])
    t3.font      = Font(bold=True, size=13, color="FFFFFF")
    t3.alignment = _center()
    ws3.row_dimensions[1].height = 28

    row = 2
    def _mem_row(ws, r, col, val, bold=False, bg=None):
        c = ws.cell(row=r, column=col, value=val)
        if bg:
            c.fill = _fill(bg)
        c.font      = Font(bold=bold, size=11)
        c.alignment = _left()
        c.border    = _border()

    if mem:
        m = mem.mem
        _hdr(ws3, row, 1, "İstatistik",  bg="sub"); _hdr(ws3, row, 2, "Değer", bg="sub"); row += 1

        tot = m.get("total_preds", 0)
        cor = m.get("correct", 0)
        acc = f"%{cor/tot*100:.1f}" if tot else "-"

        # NaN / inf koruması — ilk haftalarda adaptive değerler NaN olabilir
        import math as _math
        def _safe_f(v, default=0.0) -> float:
            """NaN veya inf ise default döner, openpyxl hücre bozulmasını önler."""
            if v is None: return default
            try:
                fv = float(v)
                return default if (_math.isnan(fv) or _math.isinf(fv)) else fv
            except (TypeError, ValueError):
                return default

        _adp = m.get("adaptive", {})
        banko_thr  = _safe_f(_adp.get("banko_threshold"),  0.65)
        double_thr = _safe_f(_adp.get("double_threshold"), 0.40)

        for label, val in [
            ("Toplam Hafta",    m.get("total_weeks", 0)),
            ("Toplam Tahmin",   tot),
            ("Doğru",           cor),
            ("Doğruluk",        acc),
            ("BANKO Eşiği",     f"{banko_thr:.3f}"),
            ("ÇİFT Eşiği",      f"{double_thr:.3f}"),
        ]:
            _mem_row(ws3, row, 1, label, bold=True, bg=C["light"])
            _mem_row(ws3, row, 2, val)
            row += 1

        row += 1
        _hdr(ws3, row, 1, "Bağlam",  bg="sub")
        _hdr(ws3, row, 2, "Doğru",   bg="sub")
        _hdr(ws3, row, 3, "Toplam",  bg="sub"); row += 1

        for ctx, d in m.get("context_acc", {}).items():
            if d.get("total", 0) >= 3:
                acc_pct = f"%{d['correct']/d['total']*100:.0f}"
                _mem_row(ws3, row, 1, ctx, bold=True, bg=C["light"])
                _mem_row(ws3, row, 2, f"{d['correct']} ({acc_pct})")
                _mem_row(ws3, row, 3, d["total"])
                row += 1

        # Son sürprizler (yeni format: dict {sezon: [...]})
        all_surprises = []
        raw_surp = m.get("surprises", {})
        if isinstance(raw_surp, dict):
            for s_list in raw_surp.values():
                if isinstance(s_list, list):
                    all_surprises.extend(s_list)
        elif isinstance(raw_surp, list):
            all_surprises = raw_surp
        surprises = sorted(all_surprises,
                           key=lambda x: x.get("confidence", 0),
                           reverse=True)[:5]
        if surprises:
            row += 1
            _hdr(ws3, row, 1, "Son Sürprizler", bg="sub"); row += 1
            for s in surprises:
                _mem_row(ws3, row, 1, s["mac"])
                _mem_row(ws3, row, 2, f"Tahmin:{s['pred']} Gerçek:{s['actual']}")
                _mem_row(ws3, row, 3, f"%{s['confidence']:.0f} güvende")
                row += 1
    else:
        ws3.cell(row=2, column=1, value="Henüz öğrenme verisi yok.")

    # ── Arşiv dosyasına kaydet ───────────────────────────────
    # Tek arşiv dosyası, her hafta yeni sayfa eklenir
    import os as _osx
    _base_dir = _osx.path.dirname(_osx.path.abspath(__file__))
    ARCHIVE_FILE = _osx.path.join(_base_dir, "..", "st_arsiv.xlsx")
    try:
        from config import ARCHIVE_FILE as _caf
        if _osx.path.exists(_caf): ARCHIVE_FILE = _caf
    except Exception: pass

    if os.path.exists(ARCHIVE_FILE):
        try:
            # Mevcut arşivi yükle
            from openpyxl import load_workbook
            archive = load_workbook(ARCHIVE_FILE)
        except (OSError, IOError, ValueError, TypeError):
            archive = openpyxl.Workbook()
            # Boş varsayılan sayfayı temizle
            if "Sheet" in archive.sheetnames:
                del archive["Sheet"]
    else:
        archive = openpyxl.Workbook()
        if "Sheet" in archive.sheetnames:
            del archive["Sheet"]

    # Sayfa adı = hafta kodu (örn: "ST41-2526" veya "ST37-2526")
    if week_id.startswith("ST"):
        sheet_name = week_id  # ST41-2526 → doğrudan kullan
    elif '-W' in week_id:
        sheet_name = f"W{week_id.split('-W')[-1]}-{datetime.now().year}"
    else:
        sheet_name = week_id

    # Aynı haftaya ait TÜM sayfaları sil (Simülasyon, Kupon Planları, Hafıza)
    to_del = [s for s in archive.sheetnames if s.startswith(sheet_name)]
    for s in to_del:
        del archive[s]

    # Yeni sayfaları arşive kopyala
    for ws in [ws1, ws2, ws3]:
        new_ws = archive.create_sheet(title=f"{sheet_name}-{ws.title}")
        for row in ws.iter_rows():
            for cell in row:
                new_cell = new_ws.cell(
                    row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font      = cell.font.copy()
                    new_cell.fill      = cell.fill.copy()
                    new_cell.border    = cell.border.copy()
                    new_cell.alignment = cell.alignment.copy()
        # Sütun genişliklerini kopyala
        for col, dim in ws.column_dimensions.items():
            new_ws.column_dimensions[col].width = dim.width
        # Satır yüksekliklerini kopyala
        for row_idx, dim in ws.row_dimensions.items():
            new_ws.row_dimensions[row_idx].height = dim.height

    # Sayfa sırasını düzenle: en yeni hafta öne
    def _sheet_sort_key(title):
        import re as _re_sk
        order = 0 if 'Simülasyon' in title else (1 if 'Kupon' in title else 2)
        m_st = _re_sk.match(r'ST(\d+)-(\d+)', title)
        if m_st: return (-int(m_st.group(2)), -int(m_st.group(1)), order)
        m_w = _re_sk.match(r'W(\d+)-(\d+)', title)
        if m_w: return (-int(m_w.group(2)), -int(m_w.group(1)), order)
        return (0, 0, 9)

    archive._sheets.sort(key=lambda ws: _sheet_sort_key(ws.title))

    archive.save(ARCHIVE_FILE)

    # Kaç hafta var
    week_count = len(set(s.rsplit('-', 1)[0] for s in archive.sheetnames))
    print(f"  Arşiv: {ARCHIVE_FILE} ({week_count} hafta)")
    return ARCHIVE_FILE


def refresh_memory_sheet(mem, week_id: str, matches: list):
    """
    Sonuçlar girildikten sonra Excel'i güncelle:
    - Simülasyon sayfasına SONUÇ ve ✓/✗ sütunları doldur
    - Hafıza sayfasını güncelle
    """
    import os
    from datetime import datetime
    # Mutlak yolu config'den al, yoksa fallback
    try:
        from config import ARCHIVE_FILE as _cfg_af
        archive_file = _cfg_af
    except Exception:
        archive_file = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "..", "st_arsiv.xlsx"
        )
    if not os.path.exists(archive_file):
        # Ek fallback: cwd
        cwd_path = os.path.join(os.getcwd(), "st_arsiv.xlsx")
        if os.path.exists(cwd_path):
            archive_file = cwd_path
        else:
            print(f"  [Excel] Dosya bulunamadı: {archive_file}")
            return

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.load_workbook(archive_file)

        FTR_REV = {"H":"1","D":"X","A":"2","0":"X"}

        # ── Simülasyon sayfasını güncelle ──────────────────
        sim_sheet = None
        # week_id: "ST41-2526" → prefix "ST41-2526"
        # eski format fallback: "ST40-2526"
        if week_id.startswith("ST"):
            _prefix = week_id  # ST41-2526 doğrudan
        else:
            _wm = re.search(r"W(\d+)", week_id)
            _ym = re.search(r"(\d{4})", week_id)
            _prefix = f"W{_wm.group(1)}-{_ym.group(1)}" if _wm and _ym else week_id
        for s in wb.sheetnames:
            if "Simülasyon" in s and _prefix in s:
                sim_sheet = s
                break

        if sim_sheet:
            ws = wb[sim_sheet]
            # Satır 3'ten başlayarak her maçı bul ve sonucu yaz
            written = 0
            for m in matches:
                actual = m.get("actual","")
                if not actual:
                    continue  # Bu maç için sonuç girilmemiş
                row = m["no"] + 2  # satır 3'ten başlıyor

                act_disp = FTR_REV.get(actual, actual)
                correct  = mem._is_correct(m)
                icon     = "✓" if correct else "✗"
                color    = "1E8449" if correct else "C0392B"  # yeşil/kırmızı

                # K sütunu = SONUÇ (11. sütun)
                c_sonuc = ws.cell(row=row, column=11, value=act_disp)
                c_sonuc.alignment = Alignment(horizontal="center", vertical="center")
                c_sonuc.font = Font(bold=True, size=11)

                # L sütunu = ✓/✗ (12. sütun)
                c_icon = ws.cell(row=row, column=12, value=icon)
                c_icon.alignment = Alignment(horizontal="center", vertical="center")
                c_icon.font = Font(bold=True, size=12, color=color)
                c_icon.fill = PatternFill("solid", fgColor="F2F3F4")

            # Özet: kaçta kaç
            total   = sum(1 for m in matches if m.get("actual"))
            correct = sum(1 for m in matches if m.get("actual") and mem._is_correct(m))
            last_row = len(matches) + 3
            # Özet satırını güncelle
            try:
                summ = ws.cell(row=last_row, column=1)
                old_val = str(summ.value or "")
                if "Sonuç:" not in old_val:
                    summ.value = (old_val + f"  |  Sonuç: {correct}/{total} "
                                  f"(%{correct/total*100:.0f})" if total else old_val)
            except (TypeError, ValueError, KeyError):
                pass

        # ── Hafıza sayfasını güncelle ──────────────────────
        mem_sheet = None
        for s in wb.sheetnames:
            if "Hafıza" in s and _prefix in s:
                mem_sheet = s
                break

        if mem_sheet:
            ws = wb[mem_sheet]
            m  = mem.mem
            tot = m.get("total_preds", 0)
            cor = m.get("correct", 0)
            acc = f"%{cor/tot*100:.1f}" if tot else "-"

            week_total   = sum(1 for x in matches if x.get("actual"))
            week_correct = sum(1 for x in matches if x.get("actual") and mem._is_correct(x))

            updates = [
                (3, 2, m.get("total_weeks", 0)),
                (4, 2, tot),
                (5, 2, cor),
                (6, 2, acc),
                (7, 2, f"{m['adaptive'].get('banko_threshold',0.65):.3f}"),
                (8, 2, f"{m['adaptive'].get('double_threshold',0.40):.3f}"),
            ]
            for r, c, v in updates:
                ws.cell(row=r, column=c, value=v)

            r = 11
            for ctx, d in m.get("context_acc", {}).items():
                if d.get("total", 0) >= 3:
                    acc_pct = f"%{d['correct']/d['total']*100:.0f}"
                    ws.cell(row=r, column=1, value=ctx)
                    ws.cell(row=r, column=2, value=f"{d['correct']} ({acc_pct})")
                    ws.cell(row=r, column=3, value=d["total"])
                    r += 1

            ws.cell(row=r+1, column=1, value=f"Bu hafta:")
            ws.cell(row=r+1, column=2,
                    value=f"{week_correct}/{week_total} "
                          f"(%{week_correct/week_total*100:.0f})" if week_total else "-")

        # ── Kupon Planları sayfasını güncelle ─────────────────
        kupon_sheet = None
        for s in wb.sheetnames:
            if "Kupon" in s and _prefix in s:
                kupon_sheet = s
                break

        if kupon_sheet:
            ws_k = wb[kupon_sheet]

            def _is_hit(label, act):
                sel = label.split()[-1] if label else ""
                return act in sel

            ws_k.cell(row=2, column=6, value="SONUÇ").font = Font(bold=True, size=10)
            for ci, lbl in [(7,"A ✓/✗"),(8,"B ✓/✗"),(9,"C ✓/✗")]:
                ws_k.cell(row=2, column=ci, value=lbl).font = Font(bold=True, size=10)

            for m in matches:
                actual = m.get("actual","")
                if not actual: continue
                no  = m.get("no", 0)
                row = no + 2
                act_disp = FTR_REV.get(actual, actual)
                ws_k.cell(row=row, column=6, value=act_disp).font = Font(bold=True)

                for ci, col_key in [(7,"A"),(8,"B"),(9,"C")]:
                    kupon_val = ""
                    for c_idx in range(3, 6):
                        hdr = str(ws_k.cell(row=2, column=c_idx).value or "")
                        if f"KUPON {col_key}" in hdr or col_key in hdr[:10]:
                            kupon_val = str(ws_k.cell(row=row, column=c_idx).value or "")
                            break
                    hit  = _is_hit(kupon_val, act_disp)
                    icon = "✓" if hit else "✗"
                    clr  = "1E8449" if hit else "C0392B"
                    c2 = ws_k.cell(row=row, column=ci, value=icon)
                    c2.font = Font(bold=True, size=12, color=clr)
                    c2.alignment = Alignment(horizontal="center")

        wb.save(archive_file)

    except (OSError, IOError, ImportError, TypeError, RuntimeError,
            KeyError, ValueError):
        pass
